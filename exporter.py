# exporter.py — генератор Excel v4 (Hybrid)
#
# АРХИТЕКТУРА:
#   Лист 2 «Смета Ремкон» — ИСТОЧНИК ИСТИНЫ:
#     - Все цены и нормы хранятся здесь
#     - Вычисления через формулы: =D10*E10, =SUM(F11:F15), =F10+H10
#     - Структура строк: ЗАГОЛОВОК ГРУППЫ → РАБОТЫ → МАТЕРИАЛЫ
#
#   Лист 1 «КП для клиента» — ССЫЛАЕТСЯ НА СМЕТУ:
#     - Все значения через ='Смета Ремкон'!F10
#     - Колонки: Мат./ед. | Мат./итого | Раб./ед. | Раб./итого | ИТОГО
#     - Работы: всегда видны (конкурентное преимущество)
#     - Материалы: свёрнуты (outline_level=1), но сумма видна в заголовке группы
#
# СМЕТА КОЛОНКИ (A-J):
#   A: маркер | B: наименование | C: ед.изм. | D: кол-во
#   E: цена мат./ед. | F: сумма мат.   (=D*E для строк, =SUM для заголовка)
#   G: цена раб./ед. | H: сумма раб.   (=D*G для строк, =SUM для заголовка)
#   I: итого (=F+H)  | J: норма/примечание
#
# КП КОЛОНКИ (A-J): те же + формулы ='Смета Ремкон'!...

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SMETA_SHEET = "Смета Ремкон"

# ─── ЦВЕТА ──────────────────────────────────
_C = {
    "hdr_bg":   "1F3864",
    "hdr_fg":   "FFFFFF",
    "sec_bg":   "2E75B6",
    "sec_fg":   "FFFFFF",
    "sub_bg":   "D9E2F3",
    "sub_fg":   "1F3864",
    "grp_bg":   "EEF4FB",   # заголовок позиции (всегда виден)
    "work_bg":  "FFFFFF",   # строка подработы
    "mat_bg":   "F5F9FF",   # строка материала (сворачиваемая)
    "mat_fg":   "555555",
    "tot_bg":   "E2EFDA",
    "extra_bg": "FFF2CC",
    "fin_bg":   "375623",
    "fin_fg":   "FFFFFF",
    "info_bg":  "EEF2F8",
}

RUB = '#,##0'
NUM = '#,##0.00'


def _fill(rgb):  return PatternFill("solid", fgColor=rgb)
def _font(bold=False, color="000000", size=9):
    return Font(name="Arial", bold=bold, color=color, size=size)
def _side():     return Side(style="thin", color="BFBFBF")
def _border():   return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
def _al(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _h(row):  return row   # alias for readability


def _set_row(ws, row, height=15):
    ws.row_dimensions[row].height = height


def _cell(ws, row, col, val="", bold=False, color="000000", size=9,
          bg=None, align="left", wrap=True, fmt=None, border=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font = _font(bold=bold, color=color, size=size)
    c.alignment = _al(align, wrap=wrap)
    if bg:
        c.fill = _fill(bg)
    if border:
        c.border = _border()
    if fmt:
        c.number_format = fmt
    return c


# ══════════════════════════════════════════════
# ЛИСТ 2 — СМЕТА РЕМКОН (источник)
# ══════════════════════════════════════════════
# Структура на позицию:
#   ROW_G  : заголовок группы (D=qty, F=SUM мат, H=SUM раб, I=F+H)
#   ROW_W1..n : подработы (D=$D$G*norm, G=price, H=D*G)
#   ROW_M1..m : материалы (D=$D$G*norm, E=price, F=D*E)
# Возвращает tracker: {item_id: {g, ws: [rows], ms: [rows]}}
def _build_smeta(ws, client, obj_name, proj_date,
                 items, qty, extra, fin):
    ws.title = SMETA_SHEET

    # ширина колонок: A B C D E F G H I J
    for col, w in [("A",6),("B",50),("C",9),("D",11),
                   ("E",14),("F",16),("G",14),("H",16),("I",18),("J",28)]:
        ws.column_dimensions[col].width = w

    r = 1
    ws.merge_cells(f"A{r}:J{r}")
    _cell(ws, r, 1,
          f"СМЕТА — {obj_name or 'объект'} · {client} · {proj_date}",
          bold=True, color=_C["hdr_fg"], size=11, bg=_C["hdr_bg"], align="center")
    _set_row(ws, r, 22)
    r += 2

    # заголовки таблицы
    hdrs = ["№", "Наименование", "Ед.", "Кол-во",
            "Мат.\nза ед., ₽", "Мат.\nитого, ₽",
            "Раб.\nза ед., ₽", "Раб.\nитого, ₽",
            "ИТОГО, ₽", "Норма / Примечание"]
    _set_row(ws, r, 38)
    for ci, h in enumerate(hdrs, 1):
        _cell(ws, r, ci, h, bold=True, color=_C["hdr_fg"], size=8,
              bg=_C["hdr_bg"], align="center")
    freeze_row = r + 1
    r += 1

    sections = list(dict.fromkeys(i["section"] for i in items))
    tracker = {}  # item_id → {g, ws:[], ms:[]}
    num = 1
    sec_g_rows = {}  # section → [group header rows] for section SUM

    for sec in sections:
        ws.merge_cells(f"A{r}:J{r}")
        _cell(ws, r, 1, sec.upper(), bold=True, color=_C["sec_fg"],
              size=10, bg=_C["sec_bg"], align="left", wrap=False)
        _set_row(ws, r, 18)
        r += 1

        sec_items = [i for i in items if i["section"] == sec]
        sec_g_rows[sec] = []
        cur_sub = None

        for it in sec_items:
            q = qty.get(it["id"], 0)
            if q == 0:
                continue

            sub = it.get("subsection", "")
            if sub and sub != cur_sub:
                cur_sub = sub
                ws.merge_cells(f"A{r}:J{r}")
                _cell(ws, r, 1, f"  {sub}", bold=True, color=_C["sub_fg"],
                      size=9, bg=_C["sub_bg"], align="left", wrap=False)
                _set_row(ws, r, 14)
                r += 1

            works = it.get("works", [])
            mats  = it.get("materials", [])
            n_w   = len(works)
            n_m   = len(mats)

            # ── предвычисляем строки ──
            g_row     = r
            w_rows    = list(range(g_row + 1, g_row + 1 + n_w))
            m_rows    = list(range(g_row + 1 + n_w, g_row + 1 + n_w + n_m))
            total_rows = 1 + n_w + n_m

            tracker[it["id"]] = {"g": g_row, "ws": w_rows, "ms": m_rows}
            sec_g_rows[sec].append(g_row)

            # ── ЗАГОЛОВОК ГРУППЫ ──
            f_formula = (f"=SUM(F{m_rows[0]}:F{m_rows[-1]})"
                         if m_rows else 0)
            h_formula = (f"=SUM(H{w_rows[0]}:H{w_rows[-1]})"
                         if w_rows else 0)
            i_formula = f"=F{g_row}+H{g_row}"

            _set_row(ws, g_row, 15)
            _cell(ws, g_row, 1, num, bold=True, bg=_C["grp_bg"], align="center", wrap=False)
            _cell(ws, g_row, 2, it["name"], bold=True, bg=_C["grp_bg"])
            _cell(ws, g_row, 3, it["unit"], bg=_C["grp_bg"], align="center")
            _cell(ws, g_row, 4, q, bold=True, bg=_C["grp_bg"], align="right", fmt=NUM)
            _cell(ws, g_row, 5, f"=IF(D{g_row}>0,F{g_row}/D{g_row},0)",
                  bg=_C["grp_bg"], align="right", fmt=RUB)
            _cell(ws, g_row, 6, f_formula, bold=True, bg=_C["grp_bg"], align="right", fmt=RUB)
            _cell(ws, g_row, 7, f"=IF(D{g_row}>0,H{g_row}/D{g_row},0)",
                  bg=_C["grp_bg"], align="right", fmt=RUB)
            _cell(ws, g_row, 8, h_formula, bold=True, bg=_C["grp_bg"], align="right", fmt=RUB)
            _cell(ws, g_row, 9, i_formula, bold=True, bg=_C["grp_bg"], align="right", fmt=RUB)
            _cell(ws, g_row, 10, f"осн. ед.: {it['unit']}", size=8,
                  color="666666", bg=_C["grp_bg"], align="left", wrap=False)
            r += 1

            # ── СТРОКИ ПОДРАБОТ ──
            for wr, w in zip(w_rows, works):
                d_f = f"=$D${g_row}*{w['norm']}"
                h_f = f"=D{wr}*G{wr}"
                _set_row(ws, wr, 13)
                _cell(ws, wr, 1, "", bg=_C["work_bg"])
                _cell(ws, wr, 2, f"  • {w['name']}", bg=_C["work_bg"])
                _cell(ws, wr, 3, w["unit"], bg=_C["work_bg"], align="center")
                _cell(ws, wr, 4, d_f, bg=_C["work_bg"], align="right", fmt=NUM)
                _cell(ws, wr, 5, "", bg=_C["work_bg"])
                _cell(ws, wr, 6, "", bg=_C["work_bg"])
                _cell(ws, wr, 7, w["price"], bg=_C["work_bg"], align="right", fmt=RUB)
                _cell(ws, wr, 8, h_f, bg=_C["work_bg"], align="right", fmt=RUB)
                _cell(ws, wr, 9, f"=H{wr}", bg=_C["work_bg"], align="right", fmt=RUB)
                _cell(ws, wr, 10, f"норма: {w['norm']} {w['unit']}/{it['unit']}",
                      size=8, color="888888", bg=_C["work_bg"], align="left", wrap=False)
                r += 1

            # ── СТРОКИ МАТЕРИАЛОВ ──
            for mr, m in zip(m_rows, mats):
                d_f = f"=$D${g_row}*{m['norm']}"
                f_f = f"=D{mr}*E{mr}"
                _set_row(ws, mr, 12)
                _cell(ws, mr, 1, "", color=_C["mat_fg"], size=8, bg=_C["mat_bg"])
                _cell(ws, mr, 2, f"  └ {m['name']}", color=_C["mat_fg"], size=8, bg=_C["mat_bg"])
                _cell(ws, mr, 3, m["unit"], color=_C["mat_fg"], size=8,
                      bg=_C["mat_bg"], align="center")
                _cell(ws, mr, 4, d_f, color=_C["mat_fg"], size=8,
                      bg=_C["mat_bg"], align="right", fmt=NUM)
                _cell(ws, mr, 5, m["price"], color=_C["mat_fg"], size=8,
                      bg=_C["mat_bg"], align="right", fmt=RUB)
                _cell(ws, mr, 6, f_f, color=_C["mat_fg"], size=8,
                      bg=_C["mat_bg"], align="right", fmt=RUB)
                _cell(ws, mr, 7, "", color=_C["mat_fg"], size=8, bg=_C["mat_bg"])
                _cell(ws, mr, 8, "", color=_C["mat_fg"], size=8, bg=_C["mat_bg"])
                _cell(ws, mr, 9, f"=F{mr}", color=_C["mat_fg"], size=8,
                      bg=_C["mat_bg"], align="right", fmt=RUB)
                _cell(ws, mr, 10, f"норма: {m['norm']} {m['unit']}/{it['unit']}",
                      size=8, color="888888", bg=_C["mat_bg"], align="left", wrap=False)
                r += 1

            num += 1

        # ── ИТОГ РАЗДЕЛА ──
        g_refs = sec_g_rows[sec]
        if g_refs:
            f_sec = "=SUM(" + "+".join(f"F{gr}" for gr in g_refs) + ")"
            h_sec = "=SUM(" + "+".join(f"H{gr}" for gr in g_refs) + ")"
            i_sec = f"=F{r}+H{r}"
            ws.merge_cells(f"A{r}:D{r}")
            _cell(ws, r, 1, f"Итого «{sec}»:", bold=True,
                  bg=_C["tot_bg"], align="right", wrap=False)
            for ci, v in [(5,""), (6, f_sec), (7,""), (8, h_sec), (9, f"=F{r}+H{r}")]:
                c = _cell(ws, r, ci, v, bold=True, bg=_C["tot_bg"], align="right",
                          fmt=RUB if ci in (6,8,9) and isinstance(v,str) and v.startswith("=") else None)
                if ci in (6, 8, 9) and isinstance(v, str) and v.startswith("="):
                    c.number_format = RUB
            _cell(ws, r, 10, "", bg=_C["tot_bg"])
            _set_row(ws, r, 16)
        r += 1

    # ── Дополнительные расходы ──
    r += 1
    ws.merge_cells(f"A{r}:J{r}")
    _cell(ws, r, 1, "ДОПОЛНИТЕЛЬНЫЕ РАСХОДЫ", bold=True, color="FFFFFF",
          size=10, bg="BF8F00", align="left", wrap=False)
    _set_row(ws, r, 18)
    extra_rows = []
    r += 1
    for name, amt in extra.items():
        if amt and amt > 0:
            ws.merge_cells(f"A{r}:H{r}")
            _cell(ws, r, 1, name, bg=_C["extra_bg"], align="left", wrap=False)
            _cell(ws, r, 9, int(amt), bold=True, bg=_C["extra_bg"], align="right", fmt=RUB)
            _cell(ws, r, 10, "", bg=_C["extra_bg"])
            extra_rows.append(r)
            _set_row(ws, r, 15)
            r += 1

    # ── Финансовый итог ──
    r += 1
    # Собираем все группы
    all_g = [tracker[it["id"]]["g"] for it in items if it["id"] in tracker]
    if all_g:
        gw_sum = "=SUM(" + "+".join(f"H{g}" for g in all_g) + ")"
        gm_sum = "=SUM(" + "+".join(f"F{g}" for g in all_g) + ")"
    else:
        gw_sum, gm_sum = 0, 0
    e_sum = ("=SUM(" + "+".join(f"I{er}" for er in extra_rows) + ")"
             if extra_rows else 0)

    fin_block = [
        ("Итого работы:", gw_sum,          _C["tot_bg"],  "000000"),
        ("Итого материалы:", gm_sum,        _C["tot_bg"],  "000000"),
        ("Итого работы + материалы:", f"=I{r}+I{r+1}", _C["tot_bg"], "000000"),
        ("Дополнительные расходы:", e_sum,  _C["extra_bg"],"000000"),
    ]
    overhead_row = profit_row = base_row = None
    for label, val, bg, fg in fin_block:
        ws.merge_cells(f"A{r}:H{r}")
        _cell(ws, r, 1, label, bold=True, color=fg, size=10, bg=bg,
              align="right", wrap=False)
        c = _cell(ws, r, 9, val, bold=True, color=fg, size=10, bg=bg, align="right")
        if isinstance(val, str) and val.startswith("="):
            c.number_format = RUB
        elif isinstance(val, (int, float)):
            c.number_format = RUB
        _cell(ws, r, 10, "", bg=bg)
        _set_row(ws, r, 17)
        if label.startswith("Итого работы + материалы"):
            wm_row = r
        r += 1

    if fin["overhead_pct"] > 0:
        overhead_row = r
        ws.merge_cells(f"A{r}:H{r}")
        _cell(ws, r, 1, f"Накладные расходы ({fin['overhead_pct']}%):",
              bold=True, size=10, bg=_C["info_bg"], align="right", wrap=False)
        c = _cell(ws, r, 9, f"=I{wm_row}*{fin['overhead_pct']}/100",
                  bold=True, size=10, bg=_C["info_bg"], align="right")
        c.number_format = RUB
        _cell(ws, r, 10, "", bg=_C["info_bg"])
        _set_row(ws, r, 17)
        r += 1

    if fin["profit_pct"] > 0:
        profit_row = r
        ws.merge_cells(f"A{r}:H{r}")
        _cell(ws, r, 1, f"Прибыль ({fin['profit_pct']}%):",
              bold=True, size=10, bg=_C["info_bg"], align="right", wrap=False)
        c = _cell(ws, r, 9, f"=I{wm_row}*{fin['profit_pct']}/100",
                  bold=True, size=10, bg=_C["info_bg"], align="right")
        c.number_format = RUB
        _cell(ws, r, 10, "", bg=_C["info_bg"])
        _set_row(ws, r, 17)
        r += 1

    # Итого без НДС
    parts = [f"I{wm_row}", f"I{wm_row-1}"]  # wm + extras
    if overhead_row: parts.append(f"I{overhead_row}")
    if profit_row:   parts.append(f"I{profit_row}")
    base_row = r
    ws.merge_cells(f"A{r}:H{r}")
    _cell(ws, r, 1, "Итого без НДС:", bold=True, size=10, bg="D9EAD3", align="right", wrap=False)
    c = _cell(ws, r, 9, "=SUM(" + "+".join(parts) + ")",
              bold=True, size=10, bg="D9EAD3", align="right")
    c.number_format = RUB
    _cell(ws, r, 10, "", bg="D9EAD3")
    _set_row(ws, r, 17)
    r += 1

    if fin["vat"]:
        vat_row = r
        ws.merge_cells(f"A{r}:H{r}")
        _cell(ws, r, 1, "НДС 20%:", bold=True, size=10, bg=_C["info_bg"], align="right", wrap=False)
        c = _cell(ws, r, 9, f"=I{base_row}*0.20",
                  bold=True, size=10, bg=_C["info_bg"], align="right")
        c.number_format = RUB
        _cell(ws, r, 10, "", bg=_C["info_bg"])
        _set_row(ws, r, 17)
        r += 1
        final_row = r
        ws.merge_cells(f"A{r}:H{r}")
        _cell(ws, r, 1, "ИТОГО С НДС:", bold=True, color=_C["fin_fg"],
              size=11, bg=_C["fin_bg"], align="right", wrap=False)
        c = _cell(ws, r, 9, f"=I{base_row}+I{vat_row}",
                  bold=True, color=_C["fin_fg"], size=11, bg=_C["fin_bg"], align="right")
        c.number_format = RUB
        _cell(ws, r, 10, "", bg=_C["fin_bg"])
        _set_row(ws, r, 22)
    else:
        final_row = r
        ws.merge_cells(f"A{r}:H{r}")
        _cell(ws, r, 1, "ИТОГО К ОПЛАТЕ:", bold=True, color=_C["fin_fg"],
              size=11, bg=_C["fin_bg"], align="right", wrap=False)
        c = _cell(ws, r, 9, f"=I{base_row}",
                  bold=True, color=_C["fin_fg"], size=11, bg=_C["fin_bg"], align="right")
        c.number_format = RUB
        _cell(ws, r, 10, "", bg=_C["fin_bg"])
        _set_row(ws, r, 22)

    ws.freeze_panes = f"A{freeze_row}"
    return tracker, final_row, base_row


# ══════════════════════════════════════════════
# ЛИСТ 1 — КП ДЛЯ КЛИЕНТА (ссылается на Смету)
# ══════════════════════════════════════════════
# Структура на позицию:
#   ROW_H  : заголовок — F=мат.итого, H=раб.итого (ссылки на Смету)
#   ROW_M1..m : строки материалов (collapsed, outline=1)
#   ROW_W1..n : строки подработ (visible)
def _build_kp(ws, client, address, obj_name, area, proj_date,
              manager, items, qty, extra, fin, tracker,
              smeta_base_row, smeta_final_row):
    ws.title = "КП для клиента"
    ws.sheet_properties.outlinePr.summaryBelow = False  # кнопка [-] над группой

    # ширина: A № | B название | C ед | D кол | E мат/ед | F мат итого | G раб/ед | H раб итого | I итого | J прим
    for col, w in [("A",6),("B",52),("C",9),("D",10),
                   ("E",15),("F",16),("G",15),("H",16),("I",18),("J",22)]:
        ws.column_dimensions[col].width = w

    def S(col, row):
        """Формула ссылки на ячейку в листе Смета"""
        return f"='{SMETA_SHEET}'!{col}{row}"

    r = 1

    # ── шапка КП ──
    ws.merge_cells(f"A{r}:J{r}")
    _cell(ws, r, 1, "КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ", bold=True,
          color=_C["hdr_fg"], size=13, bg=_C["hdr_bg"], align="center")
    ws[f"A{r}"].alignment = _al("center", "center", False)
    _set_row(ws, r, 26)
    r += 1

    meta = [("Заказчик:", client), ("Объект:", obj_name),
            ("Адрес:", address),   ("Площадь:", f"{area} м²"),
            ("Дата:", str(proj_date)), ("Менеджер:", manager)]
    for lbl, val in meta:
        ws.merge_cells(f"B{r}:J{r}")
        _cell(ws, r, 1, lbl, bold=True, bg=_C["info_bg"], border=False)
        _cell(ws, r, 2, val, bg=_C["info_bg"], border=False)
        _set_row(ws, r, 14)
        r += 1

    r += 1

    # ── заголовки таблицы ──
    hdrs = ["№ п/п", "Наименование работ / материалов", "Ед.", "Кол-во",
            "Мат.\nза ед., ₽", "Мат.\nитого, ₽",
            "Работы\nза ед., ₽", "Работы\nитого, ₽",
            "ИТОГО\n₽", "Примечание"]
    _set_row(ws, r, 40)
    for ci, h in enumerate(hdrs, 1):
        _cell(ws, r, ci, h, bold=True, color=_C["hdr_fg"], size=8,
              bg=_C["hdr_bg"], align="center")
    freeze_row = r + 1
    r += 1

    sections = list(dict.fromkeys(i["section"] for i in items))
    num = 1
    sec_header_g_rows_kp = {}  # section → [group header rows in KP] for section SUM

    for sec in sections:
        ws.merge_cells(f"A{r}:J{r}")
        _cell(ws, r, 1, sec.upper(), bold=True, color=_C["sec_fg"],
              size=10, bg=_C["sec_bg"], align="left", wrap=False)
        _set_row(ws, r, 18)
        r += 1

        sec_items = [i for i in items if i["section"] == sec]
        sec_header_g_rows_kp[sec] = []

        for it in sec_items:
            q = qty.get(it["id"], 0)
            if q == 0 or it["id"] not in tracker:
                continue

            tr = tracker[it["id"]]
            g_s  = tr["g"]      # group row в Смете
            ws_s = tr["ws"]     # work rows в Смете
            ms_s = tr["ms"]     # mat rows в Смете

            works = it.get("works", [])
            mats  = it.get("materials", [])

            # ── ЗАГОЛОВОК ПОЗИЦИИ (всегда виден) ──
            kp_g_row = r
            sec_header_g_rows_kp[sec].append(kp_g_row)

            _set_row(ws, kp_g_row, 15)
            _cell(ws, kp_g_row, 1, num, bold=True, bg=_C["grp_bg"], align="center", wrap=False)
            _cell(ws, kp_g_row, 2, it["name"], bold=True, bg=_C["grp_bg"])
            _cell(ws, kp_g_row, 3, S("C", g_s), bg=_C["grp_bg"], align="center", wrap=False)
            _cell(ws, kp_g_row, 4, S("D", g_s), bold=True, bg=_C["grp_bg"],
                  align="right", fmt=NUM)
            _cell(ws, kp_g_row, 5, S("E", g_s), bg=_C["grp_bg"], align="right", fmt=RUB)
            _cell(ws, kp_g_row, 6, S("F", g_s), bold=True, bg=_C["grp_bg"],
                  align="right", fmt=RUB)
            _cell(ws, kp_g_row, 7, S("G", g_s), bg=_C["grp_bg"], align="right", fmt=RUB)
            _cell(ws, kp_g_row, 8, S("H", g_s), bold=True, bg=_C["grp_bg"],
                  align="right", fmt=RUB)
            _cell(ws, kp_g_row, 9, f"=F{kp_g_row}+H{kp_g_row}",
                  bold=True, bg=_C["grp_bg"], align="right", fmt=RUB)
            _cell(ws, kp_g_row, 10, "", bg=_C["grp_bg"])
            r += 1

            # ── МАТЕРИАЛЫ (сворачиваемые, outline_level=1) ──
            for mr_s, m in zip(ms_s, mats):
                _set_row(ws, r, 12)
                _cell(ws, r, 1, "", color=_C["mat_fg"], size=8, bg=_C["mat_bg"])
                _cell(ws, r, 2, f"  └ {m['name']}", color=_C["mat_fg"], size=8, bg=_C["mat_bg"])
                _cell(ws, r, 3, S("C", mr_s), color=_C["mat_fg"], size=8,
                      bg=_C["mat_bg"], align="center", wrap=False)
                _cell(ws, r, 4, S("D", mr_s), color=_C["mat_fg"], size=8,
                      bg=_C["mat_bg"], align="right", fmt=NUM)
                _cell(ws, r, 5, S("E", mr_s), color=_C["mat_fg"], size=8,
                      bg=_C["mat_bg"], align="right", fmt=RUB)
                _cell(ws, r, 6, S("F", mr_s), color=_C["mat_fg"], size=8,
                      bg=_C["mat_bg"], align="right", fmt=RUB)
                _cell(ws, r, 7, "", color=_C["mat_fg"], size=8, bg=_C["mat_bg"])
                _cell(ws, r, 8, "", color=_C["mat_fg"], size=8, bg=_C["mat_bg"])
                _cell(ws, r, 9, S("F", mr_s), color=_C["mat_fg"], size=8,
                      bg=_C["mat_bg"], align="right", fmt=RUB)
                _cell(ws, r, 10, f"норма: {m['norm']} {m['unit']}/{it['unit']}",
                      size=8, color="888888", bg=_C["mat_bg"], align="left", wrap=False)
                ws.row_dimensions[r].outline_level = 1
                ws.row_dimensions[r].hidden = True
                r += 1

            # ── РАБОТЫ (всегда видны) ──
            for wr_s, w in zip(ws_s, works):
                _set_row(ws, r, 13)
                _cell(ws, r, 1, "", bg=_C["work_bg"])
                _cell(ws, r, 2, f"  • {w['name']}", bg=_C["work_bg"])
                _cell(ws, r, 3, S("C", wr_s), bg=_C["work_bg"], align="center", wrap=False)
                _cell(ws, r, 4, S("D", wr_s), bg=_C["work_bg"], align="right", fmt=NUM)
                _cell(ws, r, 5, "", bg=_C["work_bg"])
                _cell(ws, r, 6, "", bg=_C["work_bg"])
                _cell(ws, r, 7, S("G", wr_s), bg=_C["work_bg"], align="right", fmt=RUB)
                _cell(ws, r, 8, S("H", wr_s), bg=_C["work_bg"], align="right", fmt=RUB)
                _cell(ws, r, 9, S("H", wr_s), bg=_C["work_bg"], align="right", fmt=RUB)
                _cell(ws, r, 10, "", bg=_C["work_bg"])
                r += 1

            num += 1

        # ── итог раздела ──
        g_kp_refs = sec_header_g_rows_kp[sec]
        if g_kp_refs:
            ws.merge_cells(f"A{r}:D{r}")
            _cell(ws, r, 1, f"Итого «{sec}»:", bold=True,
                  bg=_C["tot_bg"], align="right", wrap=False)
            for ci, col_letter in [(5,"E"),(6,"F"),(7,"G"),(8,"H"),(9,"I")]:
                v = ("=SUM(" + "+".join(f"{col_letter}{gr}" for gr in g_kp_refs) + ")"
                     if col_letter in ("F","H","I") else "")
                c = _cell(ws, r, ci, v, bold=True, bg=_C["tot_bg"], align="right")
                if v and v.startswith("="):
                    c.number_format = RUB
            _cell(ws, r, 10, "", bg=_C["tot_bg"])
            _set_row(ws, r, 16)
        r += 1

    # ── Дополнительные расходы ──
    r += 1
    ws.merge_cells(f"A{r}:J{r}")
    _cell(ws, r, 1, "ДОПОЛНИТЕЛЬНЫЕ РАСХОДЫ", bold=True, color="FFFFFF",
          size=10, bg="BF8F00", align="left", wrap=False)
    _set_row(ws, r, 18)
    r += 1
    ne = 1
    for name, amt in extra.items():
        if amt and amt > 0:
            ws.merge_cells(f"A{r}:H{r}")
            _cell(ws, r, 1, f"{ne}. {name}", bg=_C["extra_bg"], align="left", wrap=False)
            _cell(ws, r, 9, int(amt), bold=True, bg=_C["extra_bg"], align="right", fmt=RUB)
            _cell(ws, r, 10, "", bg=_C["extra_bg"])
            _set_row(ws, r, 15)
            ne += 1; r += 1

    # ── Финансовый итог (ссылаемся на Смету) ──
    r += 1
    fin_meta = [
        ("Итого работы:",              S("I", smeta_base_row - (5 if fin["vat"] else 4)),
         _C["tot_bg"],  "000000"),
        ("Итого материалы:",           S("I", smeta_base_row - (4 if fin["vat"] else 3)),
         _C["tot_bg"],  "000000"),
        ("Итого работы + материалы:",  S("I", smeta_base_row - (3 if fin["vat"] else 2)),
         _C["tot_bg"],  "000000"),
        ("Дополнительные расходы:",    S("I", smeta_base_row - (2 if fin["vat"] else 1)),
         _C["extra_bg"],"000000"),
    ]
    # Проще: ссылаться на Смету с нужным смещением от base_row
    # Вместо сложного вычисления — дублируем финансовый блок напрямую из смет-листа
    # используем smeta_base_row как строку "Итого без НДС"
    fin_labels_smeta = []  # (label, smeta_row_9_value)

    # Перечислим финансовые строки явно через количество строк выше base_row
    # Смета: ..wm_row(base-4 или -3), extra(base-3 или -2), overhead?, profit?, base_row
    # Проще ссылаться прямо на ячейки Сметы
    # Мы знаем smeta_base_row = строка "Итого без НДС"
    fin_rows_kp = [
        (f"Итого без НДС:", S("I", smeta_base_row), "D9EAD3", "000000"),
    ]
    if fin["vat"]:
        fin_rows_kp.append((f"НДС 20%:", S("I", smeta_base_row + 1), _C["info_bg"], "000000"))
        fin_rows_kp.append(("ИТОГО С НДС:", S("I", smeta_final_row), _C["fin_bg"], _C["fin_fg"]))
    else:
        fin_rows_kp.append(("ИТОГО К ОПЛАТЕ:", S("I", smeta_final_row), _C["fin_bg"], _C["fin_fg"]))

    for label, val, bg, fg in fin_rows_kp:
        ws.merge_cells(f"A{r}:H{r}")
        _cell(ws, r, 1, label, bold=True, color=fg, size=10, bg=bg, align="right", wrap=False)
        c = _cell(ws, r, 9, val, bold=True, color=fg, size=10, bg=bg, align="right")
        c.number_format = RUB
        _cell(ws, r, 10, "", bg=bg)
        _set_row(ws, r, 20 if bg == _C["fin_bg"] else 17)
        r += 1

    ws.freeze_panes = f"A{freeze_row}"


# ══════════════════════════════════════════════
# ПУБЛИЧНАЯ ФУНКЦИЯ
# ══════════════════════════════════════════════
def generate_excel(client, address, obj_name, area, proj_date, manager,
                   selected_items, quantities, extra_costs, fin_settings) -> bytes:
    """
    fin_settings: {overhead_pct: int, profit_pct: int, vat: bool}
    Лист 2 (Смета) строится первым и возвращает tracker с номерами строк.
    Лист 1 (КП) ссылается на Смету через формулы ='Смета Ремкон'!X{row}.
    """
    wb = Workbook()
    ws_smeta = wb.active          # Лист 2 строится ПЕРВЫМ
    ws_kp    = wb.create_sheet()  # Лист 1 — после

    tracker, smeta_final_row, smeta_base_row = _build_smeta(
        ws_smeta, client, obj_name, proj_date,
        selected_items, quantities, extra_costs, fin_settings
    )
    _build_kp(
        ws_kp, client, address, obj_name, area, proj_date, manager,
        selected_items, quantities, extra_costs, fin_settings,
        tracker, smeta_base_row, smeta_final_row
    )

    # Переставляем листы: КП первым, Смета второй
    wb.move_sheet(ws_kp, offset=-1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
